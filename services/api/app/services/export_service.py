"""
Export Service for PDF and Excel Report Generation
รองรับการ export รายงานเป็น PDF และ Excel
"""

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class PDFExportService:
    """Service for generating PDF reports using ReportLab"""

    @staticmethod
    def format_thai_number(value: float | Decimal | None) -> str:
        """Format number with Thai formatting (comma separators)"""
        if value is None:
            return "0.00"
        return f"{float(value):,.2f}"

    @staticmethod
    def generate_profit_loss_pdf(data: Dict[str, Any]) -> BytesIO:
        """Generate Profit & Loss Statement PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12,
            alignment=1  # Center
        )
        story.append(Paragraph("งบกำไร-ขาดทุน (Profit & Loss Statement)", title_style))

        # Date range
        date_range = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        story.append(Paragraph(date_range, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Summary section
        summary = data['summary']
        summary_data = [
            ['รายได้รวม (Total Revenue)', f"฿{PDFExportService.format_thai_number(summary['total_revenue'])}"],
            ['ต้นทุนขาย (COGS)', f"฿{PDFExportService.format_thai_number(summary['total_cogs'])}"],
            ['กำไรขั้นต้น (Gross Profit)', f"฿{PDFExportService.format_thai_number(summary['gross_profit'])}"],
            ['Gross Margin %', f"{summary['gross_margin_percent']}%"],
        ]

        summary_table = Table(summary_data, colWidths=[3.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Income Statement Details
        income_statement = data['income_statement']
        story.append(Paragraph("รายละเอียด (Details)", styles['Heading2']))

        detail_data = [
            ['รายการ', 'จำนวนเงิน (฿)'],
            ['รายได้', ''],
            ['  ยอดขาย', PDFExportService.format_thai_number(income_statement['revenue']['total_sales'])],
            ['  ภาษีขาย (Output VAT)', PDFExportService.format_thai_number(income_statement['revenue']['output_vat'])],
            ['ต้นทุนขาย', ''],
            ['  ต้นทุนสินค้า', PDFExportService.format_thai_number(income_statement['cost_of_goods_sold']['total_cogs'])],
            ['  ภาษีซื้อ (Input VAT)', PDFExportService.format_thai_number(income_statement['cost_of_goods_sold']['input_vat_estimate'])],
            ['กำไรขั้นต้น', PDFExportService.format_thai_number(income_statement['gross_profit']['amount'])],
        ]

        detail_table = Table(detail_data, colWidths=[3.5*inch, 2*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 7), (-1, 7), colors.HexColor('#d4f1d4')),
            ('FONTNAME', (0, 7), (-1, 7), 'Helvetica-Bold'),
        ]))

        story.append(detail_table)

        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer_text = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))

        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_vat_sales_pdf(data: Dict[str, Any]) -> BytesIO:
        """Generate VAT Sales Report PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12,
            alignment=1
        )
        story.append(Paragraph("รายงานภาษีขาย (VAT Sales Report)", title_style))

        # Date range
        date_range = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        story.append(Paragraph(date_range, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Summary
        summary = data['summary']
        summary_data = [
            ['มูลค่าก่อน VAT', f"฿{PDFExportService.format_thai_number(summary['total_before_vat'])}"],
            ['VAT 7%', f"฿{PDFExportService.format_thai_number(summary['total_vat_amount'])}"],
            ['มูลค่ารวม VAT', f"฿{PDFExportService.format_thai_number(summary['total_including_vat'])}"],
        ]

        summary_table = Table(summary_data, colWidths=[3.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fef3c7')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Details table (first 50 rows to avoid PDF size issues)
        details = data['details'][:50]
        if details:
            story.append(Paragraph("รายละเอียด (Details)", styles['Heading2']))

            detail_data = [['เลขที่', 'วันที่', 'จำนวน', 'ก่อน VAT', 'VAT', 'รวม VAT']]
            for item in details:
                detail_data.append([
                    item['order_number'][:10],
                    item['order_date'],
                    str(item['quantity']),
                    PDFExportService.format_thai_number(item['price_before_vat']),
                    PDFExportService.format_thai_number(item['vat_amount']),
                    PDFExportService.format_thai_number(item['price_including_vat']),
                ])

            detail_table = Table(detail_data, colWidths=[1*inch, 1*inch, 0.7*inch, 1.2*inch, 1*inch, 1.3*inch])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            story.append(detail_table)

            if len(data['details']) > 50:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph(f"* แสดง 50 รายการแรก จากทั้งหมด {len(data['details'])} รายการ", styles['Italic']))

        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer_text = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))

        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_vat_purchases_pdf(data: Dict[str, Any]) -> BytesIO:
        """Generate VAT Purchases Report PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12,
            alignment=1
        )
        story.append(Paragraph("รายงานภาษีซื้อ (VAT Purchases Report)", title_style))

        # Date range
        date_range = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        story.append(Paragraph(date_range, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Summary
        summary = data['summary']
        summary_data = [
            ['มูลค่าก่อน VAT', f"฿{PDFExportService.format_thai_number(summary['total_before_vat'])}"],
            ['VAT ที่ซื้อ', f"฿{PDFExportService.format_thai_number(summary['total_vat_amount'])}"],
            ['มูลค่ารวม VAT', f"฿{PDFExportService.format_thai_number(summary['total_including_vat'])}"],
        ]

        summary_table = Table(summary_data, colWidths=[3.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#d4f1d4')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Details table (first 50 rows)
        details = data['details'][:50]
        if details:
            story.append(Paragraph("รายละเอียด (Details)", styles['Heading2']))

            detail_data = [['PO #', 'วันที่', 'จำนวน', 'ก่อน VAT', 'VAT', 'รวม VAT']]
            for item in details:
                detail_data.append([
                    item['po_number'][:10],
                    item['date'],
                    str(item['quantity']),
                    PDFExportService.format_thai_number(item['line_before_vat']),
                    PDFExportService.format_thai_number(item['line_vat']),
                    PDFExportService.format_thai_number(item['line_total']),
                ])

            detail_table = Table(detail_data, colWidths=[1*inch, 1*inch, 0.7*inch, 1.2*inch, 1*inch, 1.3*inch])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            story.append(detail_table)

            if len(data['details']) > 50:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph(f"* แสดง 50 รายการแรก จากทั้งหมด {len(data['details'])} รายการ", styles['Italic']))

        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer_text = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))

        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_cogs_pdf(data: Dict[str, Any]) -> BytesIO:
        """Generate COGS Report PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12,
            alignment=1
        )
        story.append(Paragraph("รายงานต้นทุนขาย (COGS Report)", title_style))

        # Date range
        date_range = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        story.append(Paragraph(date_range, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Summary
        summary = data['summary']
        summary_data = [
            ['รายได้รวม', f"฿{PDFExportService.format_thai_number(summary['total_revenue'])}"],
            ['ต้นทุนขาย', f"฿{PDFExportService.format_thai_number(summary['total_cogs'])}"],
            ['กำไรขั้นต้น', f"฿{PDFExportService.format_thai_number(summary['gross_profit'])}"],
            ['Gross Margin %', f"{summary['gross_margin_percent']}%"],
            ['จำนวนที่ขาย', f"{summary['total_quantity_sold']} ชิ้น"],
        ]

        summary_table = Table(summary_data, colWidths=[3.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Details table (first 30 rows for COGS - has more columns)
        details = data['details'][:30]
        if details:
            story.append(Paragraph("รายละเอียด (Details)", styles['Heading2']))

            detail_data = [['Order#', 'สินค้า', 'จำนวน', 'ต้นทุน/หน่วย', 'ต้นทุนรวม', 'รายได้', 'กำไร']]
            for item in details:
                detail_data.append([
                    item['order_number'][:8],
                    item['product_name'][:15],
                    str(item['quantity']),
                    PDFExportService.format_thai_number(item['unit_cost']),
                    PDFExportService.format_thai_number(item['total_cost']),
                    PDFExportService.format_thai_number(item['revenue']),
                    PDFExportService.format_thai_number(item['profit']),
                ])

            detail_table = Table(detail_data, colWidths=[0.8*inch, 1.5*inch, 0.6*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            story.append(detail_table)

            if len(data['details']) > 30:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph(f"* แสดง 30 รายการแรก จากทั้งหมด {len(data['details'])} รายการ", styles['Italic']))

        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer_text = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))

        doc.build(story)
        buffer.seek(0)
        return buffer


class ExcelExportService:
    """Service for generating Excel reports using openpyxl"""

    @staticmethod
    def format_currency(value: float | Decimal | None) -> float:
        """Format value as float for Excel"""
        if value is None:
            return 0.0
        return float(value)

    @staticmethod
    def style_header_row(ws, row_num: int, columns: int):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col in range(1, columns + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    @staticmethod
    def generate_profit_loss_excel(data: Dict[str, Any]) -> BytesIO:
        """Generate Profit & Loss Statement Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"

        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "งบกำไร-ขาดทุน (Profit & Loss Statement)"
        title_cell.font = Font(bold=True, size=16, color="1E3A8A")
        title_cell.alignment = Alignment(horizontal='center')

        # Date range
        ws.merge_cells('A2:D2')
        date_cell = ws['A2']
        date_cell.value = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        date_cell.alignment = Alignment(horizontal='center')

        # Summary section
        row = 4
        ws[f'A{row}'] = "สรุป (Summary)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        summary = data['summary']
        summary_data = [
            ('รายได้รวม (Total Revenue)', ExcelExportService.format_currency(summary['total_revenue'])),
            ('ต้นทุนขาย (COGS)', ExcelExportService.format_currency(summary['total_cogs'])),
            ('กำไรขั้นต้น (Gross Profit)', ExcelExportService.format_currency(summary['gross_profit'])),
            ('Gross Margin %', summary['gross_margin_percent']),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if label != 'Gross Margin %':
                ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Income Statement Details
        row += 2
        ws[f'A{row}'] = "รายละเอียด (Details)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        income_statement = data['income_statement']
        detail_data = [
            ('รายการ', 'จำนวนเงิน (฿)'),
            ('รายได้', ''),
            ('  ยอดขาย', ExcelExportService.format_currency(income_statement['revenue']['total_sales'])),
            ('  ภาษีขาย (Output VAT)', ExcelExportService.format_currency(income_statement['revenue']['output_vat'])),
            ('ต้นทุนขาย', ''),
            ('  ต้นทุนสินค้า', ExcelExportService.format_currency(income_statement['cost_of_goods_sold']['total_cogs'])),
            ('  ภาษีซื้อ (Input VAT)', ExcelExportService.format_currency(income_statement['cost_of_goods_sold']['input_vat_estimate'])),
            ('กำไรขั้นต้น', ExcelExportService.format_currency(income_statement['gross_profit']['amount'])),
        ]

        for i, (label, value) in enumerate(detail_data):
            ws[f'A{row + i}'] = label
            if value != '':
                ws[f'B{row + i}'] = value
                if i > 0:  # Skip header
                    ws[f'B{row + i}'].number_format = '#,##0.00'

            if i == 0:  # Header
                ExcelExportService.style_header_row(ws, row + i, 2)
            elif i == len(detail_data) - 1:  # Gross profit row
                ws[f'A{row + i}'].font = Font(bold=True)
                ws[f'B{row + i}'].font = Font(bold=True)
                ws[f'A{row + i}'].fill = PatternFill(start_color="D4F1D4", end_color="D4F1D4", fill_type="solid")
                ws[f'B{row + i}'].fill = PatternFill(start_color="D4F1D4", end_color="D4F1D4", fill_type="solid")

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        # Footer
        row += len(detail_data) + 2
        ws[f'A{row}'] = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_vat_sales_excel(data: Dict[str, Any]) -> BytesIO:
        """Generate VAT Sales Report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "VAT Sales"

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "รายงานภาษีขาย (VAT Sales Report)"
        title_cell.font = Font(bold=True, size=16, color="1E3A8A")
        title_cell.alignment = Alignment(horizontal='center')

        # Date range
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        date_cell.alignment = Alignment(horizontal='center')

        # Summary
        row = 4
        ws[f'A{row}'] = "สรุป (Summary)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        summary = data['summary']
        summary_data = [
            ('มูลค่าก่อน VAT', ExcelExportService.format_currency(summary['total_before_vat'])),
            ('VAT 7%', ExcelExportService.format_currency(summary['total_vat_amount'])),
            ('มูลค่ารวม VAT', ExcelExportService.format_currency(summary['total_including_vat'])),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            row += 1

        # Details table
        row += 2
        ws[f'A{row}'] = "รายละเอียด (Details)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = ['เลขที่ออเดอร์', 'วันที่', 'จำนวน', 'มูลค่าก่อน VAT', 'VAT', 'มูลค่ารวม VAT']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header

        ExcelExportService.style_header_row(ws, row, len(headers))
        row += 1

        # Data rows
        for item in data['details']:
            ws.cell(row=row, column=1).value = item['order_number']
            ws.cell(row=row, column=2).value = item['order_date']
            ws.cell(row=row, column=3).value = item['quantity']
            ws.cell(row=row, column=4).value = ExcelExportService.format_currency(item['price_before_vat'])
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5).value = ExcelExportService.format_currency(item['vat_amount'])
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).value = ExcelExportService.format_currency(item['price_including_vat'])
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18

        # Footer
        row += 2
        ws[f'A{row}'] = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_vat_purchases_excel(data: Dict[str, Any]) -> BytesIO:
        """Generate VAT Purchases Report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "VAT Purchases"

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "รายงานภาษีซื้อ (VAT Purchases Report)"
        title_cell.font = Font(bold=True, size=16, color="1E3A8A")
        title_cell.alignment = Alignment(horizontal='center')

        # Date range
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        date_cell.alignment = Alignment(horizontal='center')

        # Summary
        row = 4
        ws[f'A{row}'] = "สรุป (Summary)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        summary = data['summary']
        summary_data = [
            ('มูลค่าก่อน VAT', ExcelExportService.format_currency(summary['total_before_vat'])),
            ('VAT ที่ซื้อ', ExcelExportService.format_currency(summary['total_vat_amount'])),
            ('มูลค่ารวม VAT', ExcelExportService.format_currency(summary['total_including_vat'])),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="D4F1D4", end_color="D4F1D4", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="D4F1D4", end_color="D4F1D4", fill_type="solid")
            row += 1

        # Details table
        row += 2
        ws[f'A{row}'] = "รายละเอียด (Details)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = ['เลขที่ PO', 'วันที่', 'จำนวน', 'มูลค่าก่อน VAT', 'VAT', 'มูลค่ารวม VAT']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header

        ExcelExportService.style_header_row(ws, row, len(headers))
        row += 1

        # Data rows
        for item in data['details']:
            ws.cell(row=row, column=1).value = item['po_number']
            ws.cell(row=row, column=2).value = item['date']
            ws.cell(row=row, column=3).value = item['quantity']
            ws.cell(row=row, column=4).value = ExcelExportService.format_currency(item['line_before_vat'])
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5).value = ExcelExportService.format_currency(item['line_vat'])
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).value = ExcelExportService.format_currency(item['line_total'])
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18

        # Footer
        row += 2
        ws[f'A{row}'] = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_cogs_excel(data: Dict[str, Any]) -> BytesIO:
        """Generate COGS Report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "COGS Report"

        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "รายงานต้นทุนขาย (COGS Report)"
        title_cell.font = Font(bold=True, size=16, color="1E3A8A")
        title_cell.alignment = Alignment(horizontal='center')

        # Date range
        ws.merge_cells('A2:H2')
        date_cell = ws['A2']
        date_cell.value = f"ระหว่างวันที่ {data['start_date']} ถึง {data['end_date']}"
        date_cell.alignment = Alignment(horizontal='center')

        # Summary
        row = 4
        ws[f'A{row}'] = "สรุป (Summary)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        summary = data['summary']
        summary_data = [
            ('รายได้รวม', ExcelExportService.format_currency(summary['total_revenue'])),
            ('ต้นทุนขาย', ExcelExportService.format_currency(summary['total_cogs'])),
            ('กำไรขั้นต้น', ExcelExportService.format_currency(summary['gross_profit'])),
            ('Gross Margin %', summary['gross_margin_percent']),
            ('จำนวนที่ขายได้', summary['total_quantity_sold']),
        ]

        for i, (label, value) in enumerate(summary_data):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if i < 3:  # First 3 are currency
                ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            row += 1

        # Details table
        row += 2
        ws[f'A{row}'] = "รายละเอียด (Details)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = ['เลขที่ออเดอร์', 'สินค้า', 'SKU', 'จำนวน', 'ต้นทุน/หน่วย', 'ต้นทุนรวม', 'รายได้', 'กำไร', 'Margin %']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header

        ExcelExportService.style_header_row(ws, row, len(headers))
        row += 1

        # Data rows
        for item in data['details']:
            ws.cell(row=row, column=1).value = item['order_number']
            ws.cell(row=row, column=2).value = item['product_name']
            ws.cell(row=row, column=3).value = item['sku']
            ws.cell(row=row, column=4).value = item['quantity']
            ws.cell(row=row, column=5).value = ExcelExportService.format_currency(item['unit_cost'])
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).value = ExcelExportService.format_currency(item['total_cost'])
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=7).value = ExcelExportService.format_currency(item['revenue'])
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8).value = ExcelExportService.format_currency(item['profit'])
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=9).value = item['profit_margin_percent']
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12

        # Footer
        row += 2
        ws[f'A{row}'] = f"สร้างรายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
